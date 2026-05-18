"""
pages/13_Derivados_Credito_CDO.py
----------------------------------
Módulo 13: Valuación de Derivados de Crédito — CDO Tranches.
Implementa el modelo de un factor gaussiano de Hull (Options, Futures,
and Other Derivatives, 9ª ed., ecuación 25.12) con cuadratura de
Gauss-Hermite para valuar el spread de cualquier tranche de un CDO.

Orden pedagógico:
  1. Estructura del CDO y parámetros del portafolio
  2. Probabilidad de default condicional Q(t|F)
  3. Distribución binomial P(k,t|F) — modelo un factor
  4. Pérdida esperada del tranche E_j(F)
  5. Fee Leg (A + B) y Protection Leg (C)
  6. Spread del tranche = C / (A + B)
  7. Análisis de sensibilidad y comparativa de tranches
  8. Exportar reporte (datos, metodología y resultados)
"""

import io
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from scipy.stats import norm
from scipy.special import gammaln

from utils import (
    get_engine, page_header, paso_a_paso, separador,
    themed_info, themed_success, themed_warning, themed_error,
    apply_plotly_theme, plotly_theme, plotly_colors, plotly_color,
    get_current_theme,
)

# =============================================================================
# CONFIGURACIÓN
# =============================================================================
st.set_page_config(
    page_title="CDO Tranches · Calculadora Financiera",
    page_icon="🏦",
    layout="wide",
)

engine = get_engine()

# --- Estilos globales para métricas destacadas ---
math_style = "font-family: 'Times New Roman', Times, serif; font-style: italic; font-weight: normal; padding: 0 2px;"
css_titulo = "font-size: 20px; opacity: 0.85; font-weight: 500;"
css_valor = "font-size: 28px; font-weight: bold;"
css_contenedor = "display: flex; justify-content: space-between; align-items: center; width: 100%; padding: 12px 0;"
css_paso = "text-align: center; font-size: 22px; font-weight: bold; padding: 4px 0; margin: 0;"

# Variante para columnas estrechas
css_contenedor_col = "display: flex; flex-direction: column; justify-content: center; align-items: center; width: 100%; padding: 8px 0;"
css_titulo_col = "font-size: 16px; opacity: 0.85; font-weight: 500; margin-bottom: 4px; text-align: center;"
css_valor_col = "font-size: 24px; font-weight: bold; text-align: center;"

page_header(
    titulo="14. Derivados de Crédito — CDO",
    subtitulo="Modelo de un Factor Gaussiano · Gauss-Hermite · Hull (9ª ed.)",
)

# =============================================================================
# MOTOR MATEMÁTICO (autocontenido, sin depender de engine)
# =============================================================================

@st.cache_data
def _gauss_hermite_normal(M: int):
    """Nodos F_k y pesos w_k para integrar sobre N(0,1) (Hull ec. 25.12)."""
    x_nodes, w_hat = np.polynomial.hermite.hermgauss(M)
    F_nodes = x_nodes * np.sqrt(2)
    w_nodes = w_hat / np.sqrt(np.pi)
    return F_nodes, w_nodes


def _log_binom_coef(n: int, k: int) -> float:
    """log C(n,k) vía gammaln — evita overflow con n grande."""
    return gammaln(n + 1) - gammaln(k + 1) - gammaln(n - k + 1)


def _Q_condicional(t: float, h: float, rho: float, F: np.ndarray) -> np.ndarray:
    """
    Probabilidad condicional de default al tiempo t dado el factor F.
        Q(t) = 1 − exp(−h·t)
        Q(t|F) = N([N⁻¹(Q(t)) − √ρ·F] / √(1−ρ))
    """
    Qt     = 1.0 - np.exp(-h * t)
    inv_Qt = norm.ppf(Qt)
    arg    = (inv_Qt - np.sqrt(rho) * F) / np.sqrt(1.0 - rho)
    return norm.cdf(arg)


def _binom_probs_log(n: int, q: np.ndarray, k_max: int) -> np.ndarray:
    """
    P(k, t|F) para k = 0 … k_max−1.
    Usa log-probabilidades para estabilidad numérica con n grande.
    Retorna array shape (k_max, M).
    """
    eps    = 1e-300
    log_q  = np.log(np.clip(q,       eps, 1 - eps))
    log_1q = np.log(np.clip(1.0 - q, eps, 1 - eps))
    probs  = np.zeros((k_max, len(q)))
    for k in range(k_max):
        log_p    = _log_binom_coef(n, k) + k * log_q + (n - k) * log_1q
        probs[k] = np.exp(log_p)
    return probs


def _E_tranche_dado_F(t, h, rho, n, R, alpha_L, alpha_H, F_nodes):
    """
    Pérdida esperada del tranche [α_L, α_H] dado cada F_k al tiempo t.
    Retorna array (M,).
    """
    q   = _Q_condicional(t, h, rho, F_nodes)
    LGD = 1.0 - R
    n_L = alpha_L * n / LGD
    n_H = alpha_H * n / LGD
    m_L = int(np.ceil(n_L))
    m_H = int(np.ceil(n_H))
    k_max = m_H
    P   = _binom_probs_log(n, q, k_max)
    E   = P[:m_L].sum(axis=0)
    for k in range(m_L, m_H):
        frac = np.clip((alpha_H - k * LGD / n) / (alpha_H - alpha_L), 0.0, 1.0)
        E   += P[k] * frac
    return E


def valuar_tranche(h, rho, n, R, alpha_L, alpha_H, tasa_rf, periodos, F_nodes, w_nodes):
    """
    Motor completo: calcula A, B, C y spread del tranche.
    Retorna dict con todos los resultados.
    """
    taus  = [0.0] + list(periodos)
    M     = len(F_nodes)
    E_all = np.zeros((len(taus), M))
    E_all[0] = 1.0  # sin pérdidas al inicio

    for j, t in enumerate(periodos):
        E_all[j + 1] = _E_tranche_dado_F(t, h, rho, n, R, alpha_L, alpha_H, F_nodes)

    A_F = np.zeros(M)
    B_F = np.zeros(M)
    C_F = np.zeros(M)

    for j in range(1, len(taus)):
        tau_j   = taus[j]
        tau_j1  = taus[j - 1]
        delta   = tau_j - tau_j1
        tau_mid = 0.5 * (tau_j + tau_j1)
        v_j     = np.exp(-tasa_rf * tau_j)
        v_mid   = np.exp(-tasa_rf * tau_mid)
        E_j     = E_all[j]
        E_j1    = E_all[j - 1]
        dE      = E_j1 - E_j
        A_F    += delta * E_j  * v_j
        B_F    += 0.5 * delta  * dE * v_mid
        C_F    += dE * v_mid

    A      = (w_nodes * A_F).sum()
    B      = (w_nodes * B_F).sum()
    C      = (w_nodes * C_F).sum()
    spread = C / (A + B) if (A + B) > 1e-15 else np.nan
    E_T    = (w_nodes * E_all[-1]).sum()

    return dict(A=A, B=B, C=C,
                spread=spread,
                spread_bps=spread * 10_000,
                E_T=E_T,
                E_all=E_all)


# Tranches predefinidos del ejemplo de Hull
TRANCHES_DEFAULT = [
    ("0–3%",    0.00, 0.03),
    ("3–6%",    0.03, 0.06),
    ("6–9%",    0.06, 0.09),
    ("9–12%",   0.09, 0.12),
    ("12–22%",  0.12, 0.22),
    ("22–100%", 0.22, 1.00),
]

TRANCHE_COLORES = {
    "0–3%":    "#2196F3",
    "3–6%":    "#4CAF50",
    "6–9%":    "#FF5722",
    "9–12%":   "#9C27B0",
    "12–22%":  "#FFC107",
    "22–100%": "#607D8B",
}

# =============================================================================
# PESTAÑAS
# =============================================================================
tab_inputs, tab_motor, tab_resultado, tab_sens, tab_todos, tab_export = st.tabs([
    "Parámetros",
    "Motor matemático",
    "Resultado",
    "Sensibilidad",
    "Todos los Tranches",
    "Exportar Reporte",
])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — PARÁMETROS
# ─────────────────────────────────────────────────────────────────────────────
with tab_inputs:
    st.markdown("### Estructura del CDO y parámetros del portafolio")
    themed_info(
        "Un **CDO (Collateralized Debt Obligation)** divide las pérdidas de un portafolio crediticio "
        "en **tranches** con distinto nivel de subordinación. El tranche más junior (equity) absorbe "
        "las primeras pérdidas; los tranches senior solo pierden cuando el portafolio colapsa masivamente. "
        "El modelo de Hull usa un **factor sistemático gaussiano** (ρ) para capturar la correlación "
        "entre defaults y valúa el spread justo de cada tranche."
    )

    separador()

    # ── Diagrama conceptual de tranches ──────────────────────────────────────
    st.markdown("#### Estructura de Tranches — Portafolio de ejemplo")
    tranche_data = {
        "Tranche":           ["22–100%", "12–22%", "9–12%", "6–9%", "3–6%", "0–3%"],
        "Attachment (%)":    [22, 12, 9, 6, 3, 0],
        "Detachment (%)":    [100, 22, 12, 9, 6, 3],
        "Ancho (%)":         [78, 10, 3, 3, 3, 3],
        "Tipo":              ["Senior", "Mezzanine", "Mezzanine", "Mezzanine", "Junior", "Equity"],
    }
    df_struct = pd.DataFrame(tranche_data)
    st.dataframe(df_struct, use_container_width=True, hide_index=True)

    separador()

    # ── Inputs del usuario ───────────────────────────────────────────────────
    st.markdown("### Parámetros del modelo")
    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("**Portafolio de referencia**")
        val_port = st.number_input(
            "Valor del portafolio ($)",
            min_value=1_000_000.0, value=62_500_000.0,
            step=500_000.0, format="%.0f", key="cdo_val_port",
        )
        n_ent = st.number_input(
            "Número de entidades de referencia (n)",
            min_value=2, max_value=10_000, value=125, step=1, key="cdo_n",
        )
        T_cds = st.number_input(
            "Vencimiento CDS (años)",
            min_value=0.25, max_value=30.0, value=1.0, step=0.25, key="cdo_T",
        )
        dt_pago = st.number_input(
            "Periodo de pago de primas (años)",
            min_value=0.0, max_value=1.0, value=0.25, step=0.25,
            format="%.4f", key="cdo_dt",
        )

    with c2:
        st.markdown("**Parámetros de crédito**")
        h_rate = st.number_input(
            "Hazard rate (h) %",
            min_value=0.001, max_value=100.0, value=0.323, step=0.01,
            format="%.4f", key="cdo_h",
        ) / 100
        RR = st.number_input(
            "Recovery Rate (RR) %",
            min_value=0.0, max_value=100.0, value=40.0, step=1.0, key="cdo_RR",
        ) / 100
        LGD = 1.0 - RR
        st.metric("LGD = 1 − RR", f"{LGD:.2%}")
        rho = st.number_input(
            "Correlación (ρ) %",
            min_value=0.0, max_value=99.0, value=25.0, step=1.0, key="cdo_rho",
        ) / 100

    with c3:
        st.markdown("**Parámetros derivados del tranche**")
        M_quad = st.slider(
            "Puntos de cuadratura Gauss-Hermite (M)",
            min_value=10, max_value=100, value=100, step=5,
            key="cdo_M",
            help="Mayor M = mayor precision. M=100 es el estandar de Hull. M>=50 ya da buena convergencia.",
        )
        rf = st.number_input(
            "Tasa libre de riesgo (r) %",
            min_value=0.0, max_value=30.0, value=3.0, step=0.1, key="cdo_rf",
        ) / 100

        # ── Selección del tranche con sincronización automática ───────────────
        st.markdown("**Selección del tranche a valuar**")
        tranche_names = [t[0] for t in TRANCHES_DEFAULT]
        tranche_sel   = st.selectbox(
            "Tranche:", tranche_names, index=1, key="cdo_tranche",
        )

        # Sincronizar α_L y α_H con el tranche seleccionado
        _sel_idx      = tranche_names.index(tranche_sel)
        _default_aL   = TRANCHES_DEFAULT[_sel_idx][1] * 100   # en %
        _default_aH   = TRANCHES_DEFAULT[_sel_idx][2] * 100   # en %

        # Si el tranche cambió, actualizar session_state antes de renderizar inputs
        if st.session_state.get("_ultimo_tranche") != tranche_sel:
            st.session_state["cdo_aL"] = _default_aL
            st.session_state["cdo_aH"] = _default_aH
            st.session_state["_ultimo_tranche"] = tranche_sel

        alpha_L_pct = st.number_input(
            "Attachment point α_L (%)",
            min_value=0.0, max_value=99.0,
            value=st.session_state.get("cdo_aL", _default_aL),
            step=0.5, key="cdo_aL",
        )
        alpha_H_pct = st.number_input(
            "Detachment point α_H (%)",
            min_value=0.01, max_value=100.0,
            value=st.session_state.get("cdo_aH", _default_aH),
            step=0.5, key="cdo_aH",
        )
        alpha_L = alpha_L_pct / 100
        alpha_H = alpha_H_pct / 100

    separador()

    # ── Parámetros derivados ─────────────────────────────────────────────────
    st.markdown("### Parámetros derivados del tranche")

    n_L = alpha_L * n_ent / LGD if LGD > 0 else 0
    n_H = alpha_H * n_ent / LGD if LGD > 0 else 0
    m_nL = int(np.ceil(n_L))
    m_nH = int(np.ceil(n_H))
    periodos = list(np.arange(dt_pago, T_cds + 1e-9, dt_pago))
    periodos = [round(p, 6) for p in periodos if p <= T_cds + 1e-9]
    # Valor nominal por tranche
    val_tranche = (alpha_H - alpha_L) * val_port
    val_tranche_acum = alpha_H * val_port

    dc1, dc2, dc3, dc4, dc5, dc6 = st.columns(6)
    
    with dc1:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'><span style='{math_style}'>α<sub>L</sub></span> (Attachment)</span>"
            f"<span style='{css_valor_col}'>{alpha_L:.2%}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )
    with dc2:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'><span style='{math_style}'>α<sub>H</sub></span> (Detachment)</span>"
            f"<span style='{css_valor_col}'>{alpha_H:.2%}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )
    with dc3:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'><span style='{math_style}'>n<sub>L</sub> = α<sub>L</sub> &middot; n</span> / LGD</span>"
            f"<span style='{css_valor_col}'>{n_L:.2f}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )
    with dc4:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'><span style='{math_style}'>n<sub>H</sub> = α<sub>H</sub> &middot; n</span> / LGD</span>"
            f"<span style='{css_valor_col}'>{n_H:.2f}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )
    with dc5:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'><span style='{math_style}'>m(n<sub>L</sub>) = &lceil;n<sub>L</sub>&rceil;</span></span>"
            f"<span style='{css_valor_col}'>{m_nL}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )
    with dc6:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'><span style='{math_style}'>m(n<sub>H</sub>) = &lceil;n<sub>H</sub>&rceil;</span></span>"
            f"<span style='{css_valor_col}'>{m_nH}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )

    separador()

    # También actualizamos la fila inferior para mantener la coherencia visual
    dv1, dv2, dv3 = st.columns(3)
    
    periodos_limpios = ", ".join([f"{float(p):.2f}" for p in periodos])
    if len(periodos_limpios) > 40:
        periodos_limpios = periodos_limpios[:35] + "..."

    with dv1:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'>Valor inicial del tranche</span>"
            f"<span style='{css_valor_col}'>${val_tranche:,.0f}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )
    with dv2:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'>Valor acumulado del tranche</span>"
            f"<span style='{css_valor_col}'>${val_tranche_acum:,.0f}</span>"
            f"</div>", 
            unsafe_allow_html=True
        )
    with dv3:
        st.markdown(
            f"<div style='{css_contenedor_col}'>"
            f"<span style='{css_titulo_col}'>Periodos de pago</span>"
            f"<span style='{css_valor_col}'>{len(periodos)} <span style='font-size: 16px; font-weight: normal; opacity: 0.8;'>({periodos_limpios})</span></span>"
            f"</div>", 
            unsafe_allow_html=True
        )

    with paso_a_paso():
        st.latex(r"n_L = \frac{\alpha_L \cdot n}{1 - R}, \quad n_H = \frac{\alpha_H \cdot n}{1 - R}")
        st.latex(
            rf"n_L = \frac{{{alpha_L:.4f} \times {n_ent}}}{{{LGD:.4f}}} = {n_L:.4f}"
            + r"\quad\Rightarrow\quad m(n_L) = \lceil n_L \rceil = " + str(m_nL)
        )
        st.latex(
            rf"n_H = \frac{{{alpha_H:.4f} \times {n_ent}}}{{{LGD:.4f}}} = {n_H:.4f}"
            + r"\quad\Rightarrow\quad m(n_H) = \lceil n_H \rceil = " + str(m_nH)
        )


# ─────────────────────────────────────────────────────────────────────────────
# CÁLCULO CENTRAL (se ejecuta antes de tabs 2-5)
# ─────────────────────────────────────────────────────────────────────────────
_h     = st.session_state.get("cdo_h",      0.323) / 100
_RR    = st.session_state.get("cdo_RR",     40.0)  / 100
_LGD   = 1.0 - _RR
_rho   = st.session_state.get("cdo_rho",    25.0)  / 100
_n     = int(st.session_state.get("cdo_n",  125))
_T     = st.session_state.get("cdo_T",      1.0)
_dt    = st.session_state.get("cdo_dt",     0.25)
_rf    = st.session_state.get("cdo_rf",     3.0)   / 100
_aL    = st.session_state.get("cdo_aL",     3.0)   / 100
_aH    = st.session_state.get("cdo_aH",     6.0)   / 100
_VP    = st.session_state.get("cdo_val_port", 62_500_000.0)
_M     = int(st.session_state.get("cdo_M",  100))

_periodos = list(np.arange(_dt, _T + 1e-9, _dt))
_periodos = [round(p, 6) for p in _periodos if p <= _T + 1e-9]

F_NODES, W_NODES = _gauss_hermite_normal(_M)

_res = valuar_tranche(
    h=_h, rho=_rho, n=_n, R=_RR,
    alpha_L=_aL, alpha_H=_aH,
    tasa_rf=_rf, periodos=_periodos,
    F_nodes=F_NODES, w_nodes=W_NODES,
)

_tranche_nombre = st.session_state.get("cdo_tranche", "3–6%")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — MOTOR MATEMÁTICO
# ─────────────────────────────────────────────────────────────────────────────
with tab_motor:
    st.markdown("### Motor: Modelo de un Factor Gaussiano (Hull, ec. 25.12)")

    themed_info(
        f"La integral sobre el factor sistémico **F** (variable latente que representa el "
        f"estado macroeconómico) se aproxima con **cuadratura de Gauss-Hermite** (M=**{_M}** nodos). "
        "Para cada valor de F se calcula la distribución binomial condicional de defaults y "
        "se obtiene la pérdida esperada del tranche."
    )

    separador()

    st.markdown("#### Paso 1 — Probabilidad acumulada de default Q(t)")
    themed_info(
        "Con un proceso de Poisson de intensidad **h** (hazard rate), "
        "la probabilidad de default al tiempo t es Q(t) = 1 − exp(−h·t)."
    )

    with paso_a_paso():
        st.latex(r"Q(t) = 1 - e^{-h \cdot t}")
        for t_j in _periodos:
            Qt_j = 1.0 - np.exp(-_h * t_j)
            st.latex(
                rf"Q({t_j:.2f}) = 1 - e^{{-{_h:.6f} \times {t_j:.2f}}} = {Qt_j:.6f} \; ({Qt_j:.4%})"
            )

    separador()

    st.markdown("#### Paso 2 — Probabilidad condicional de default Q(t|F)")
    themed_info(
        "Dado el factor sistemático **F**, la correlación ρ induce dependencia entre defaults. "
        "Con mayor ρ los defaults se agrupan (riesgo de cola aumenta)."
    )

    with paso_a_paso():
        st.latex(
            r"Q(t|F) = N\!\left(\frac{N^{-1}[Q(t)] - \sqrt{\rho}\,F}{\sqrt{1-\rho}}\right)"
        )
        t_ej = _periodos[-1]
        Qt_ej = 1.0 - np.exp(-_h * t_ej)
        inv_Qt = norm.ppf(Qt_ej)
        arg_0 = (inv_Qt - np.sqrt(_rho) * 0) / np.sqrt(1 - _rho)
        Q_0   = norm.cdf(arg_0)
        st.latex(
            rf"t = {t_ej:.2f}, \; Q(t) = {Qt_ej:.6f}, \;"
            rf"N^{{-1}}[Q(t)] = {inv_Qt:.6f}"
        )
        st.latex(
            rf"Q(t|F=0) = N\!\left(\frac{{{inv_Qt:.6f} - \sqrt{{{_rho:.4f}}}\cdot 0}}{{\sqrt{{1 - {_rho:.4f}}}}}\right) = N({arg_0:.6f}) = {Q_0:.8f}"
        )

    F_plot = np.linspace(-4, 4, 300)
    Q_plot = _Q_condicional(_periodos[-1], _h, _rho, F_plot)
    fig_q = go.Figure()
    fig_q.add_trace(go.Scatter(
        x=F_plot, y=Q_plot * 100,
        mode="lines",
        line=dict(color=plotly_colors()[0], width=2.5),
        name=f"Q(t={_periodos[-1]:.2f}|F)",
    ))
    fig_q.add_vline(x=0, line_dash="dash", line_color="gray", line_width=1,
                    annotation_text="F=0 (promedio)")
    fig_q = apply_plotly_theme(fig_q)
    fig_q.update_layout(
        **plotly_theme(),
        title="Probabilidad condicional de default Q(t|F) al vencimiento",
        xaxis_title="Factor sistemático F",
        yaxis_title="Q(t|F) (%)",
        height=360,
    )
    st.plotly_chart(fig_q, use_container_width=True)

    separador()

    st.markdown("#### Paso 3 — Distribución binomial condicional P(k, t|F)")
    themed_info(
        "Dado F, los defaults son i.i.d. con probabilidad Q(t|F). El número de defaults "
        "sigue una **binomial(n, Q(t|F))**. Para n grande se usan log-probabilidades "
        "via `gammaln` para evitar desbordamiento numérico."
    )

    with paso_a_paso():
        st.latex(
            r"P(k, t|F) = \binom{n}{k}\, Q(t|F)^k \,[1-Q(t|F)]^{n-k}"
        )
        st.latex(
            r"\log P(k,t|F) = \log\binom{n}{k} + k\log Q + (n-k)\log(1-Q)"
        )

    separador()

    st.markdown("#### Paso 4 — Pérdida esperada del tranche E_j(F)")
    themed_info(
        f"El tranche **[{_aL:.0%}, {_aH:.0%}]** empieza a perder con m(n_L) defaults y queda "
        "destruido completamente con m(n_H) defaults. Entre ambos extremos la "
        "pérdida es proporcional al número de defaults."
    )

    with paso_a_paso():
        st.latex(
            r"E_j(F) = \underbrace{\sum_{k=0}^{m(n_L)-1} P(k,\tau_j|F)\cdot 1}_{\text{tranche lleno}}"
            r"+ \underbrace{\sum_{k=m(n_L)}^{m(n_H)-1} P(k,\tau_j|F)\cdot\frac{\alpha_H - k(1-R)/n}{\alpha_H - \alpha_L}}_{\text{tranche parcial}}"
        )

    E_tab = []
    for j, t_j in enumerate([0.0] + _periodos):
        if j == 0:
            E_tab.append({"τ": "0 (inicio)", "E_j(F=0)": 1.0})
        else:
            Ej = _E_tranche_dado_F(t_j, _h, _rho, _n, _RR, _aL, _aH, np.array([0.0]))[0]
            E_tab.append({"τ": f"{t_j:.2f}", "E_j(F=0)": Ej})
    df_Ej = pd.DataFrame(E_tab)
    st.dataframe(df_Ej.style.format({"E_j(F=0)": "{:.8f}"}), use_container_width=True, hide_index=True)

    separador()

    st.markdown("#### Paso 5 — Cuadratura de Gauss-Hermite")
    themed_info(
        "Para integrar sobre el factor F ~ N(0,1) se usan M=100 nodos de cuadratura. "
        "Cada nodo tiene un peso w_k. La integral se aproxima como Σ w_k · g(F_k)."
    )

    with paso_a_paso():
        st.latex(
            r"\int_{-\infty}^{+\infty} \frac{1}{\sqrt{2\pi}} e^{-F^2/2} g(F)\,dF "
            r"\approx \sum_{k=1}^{M} w_k\, g(F_k)"
        )
        st.latex(rf"M = {_M}, \quad \sum_k w_k = {W_NODES.sum():.8f} \approx 1")

    fig_gh = go.Figure()
    fig_gh.add_trace(go.Bar(
        x=F_NODES, y=W_NODES,
        marker_color=plotly_colors()[1], opacity=0.75,
        name="Pesos w_k",
    ))
    fig_gh = apply_plotly_theme(fig_gh)
    fig_gh.update_layout(
        **plotly_theme(),
        title=f"Nodos y pesos de Gauss-Hermite (M={_M})",
        xaxis_title="F_k",
        yaxis_title="w_k",
        height=300,
    )
    st.plotly_chart(fig_gh, use_container_width=True)

    separador()

    st.markdown("#### Paso 6 — Fee Leg (A, B) y Protection Leg (C)")
    themed_info(
        "El **spread** del tranche es la prima s tal que el valor presente de los pagos de "
        "protección (C) iguala el valor presente de la fee leg (A + B)."
    )

    with paso_a_paso():
        st.latex(
            r"A(F) = \sum_{j=1}^{m}(\tau_j - \tau_{j-1})\cdot E_j(F)\cdot v(\tau_j)"
        )
        st.latex(
            r"B(F) = \sum_{j=1}^{m}\tfrac{1}{2}(\tau_j-\tau_{j-1})\cdot[E_{j-1}(F)-E_j(F)]\cdot v\!\left(\tfrac{\tau_j+\tau_{j-1}}{2}\right)"
        )
        st.latex(
            r"C(F) = \sum_{j=1}^{m}[E_{j-1}(F)-E_j(F)]\cdot v\!\left(\tfrac{\tau_j+\tau_{j-1}}{2}\right)"
        )
        st.latex(
            r"A = \sum_k w_k A(F_k), \quad B = \sum_k w_k B(F_k), \quad C = \sum_k w_k C(F_k)"
        )
        st.latex(r"s = \frac{C}{A + B}")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — RESULTADO
# ─────────────────────────────────────────────────────────────────────────────
with tab_resultado:
    st.markdown(f"### Resultado: Spread del tranche **{_tranche_nombre}**")

    spread_bps = _res["spread_bps"]
    A, B, C    = _res["A"], _res["B"], _res["C"]
    E_T        = _res["E_T"]

    themed_success(
        f"<div style='{css_contenedor}'>"
        f"<span style='{css_titulo}'>Spread del tranche <b>{_tranche_nombre}</b> "
        f"(<span style='{math_style}'>s</span>)</span>"
        f"<span style='{css_valor}'>{spread_bps:.1f} bps</span>"
        f"</div>"
    )

    # REFACTORIZADO: Uso de los estilos de columnas para reemplazar st.metric
    c1r, c2r, c3r, c4r = st.columns(4)
    with c1r:
        st.markdown(f"<div style='{css_contenedor_col}'><span style='{css_titulo_col}'>A (Fee Leg principal)</span><span style='{css_valor_col}'>{A:.6f}</span></div>", unsafe_allow_html=True)
    with c2r:
        st.markdown(f"<div style='{css_contenedor_col}'><span style='{css_titulo_col}'>B (Fee Leg accrual)</span><span style='{css_valor_col}'>{B:.6f}</span></div>", unsafe_allow_html=True)
    with c3r:
        st.markdown(f"<div style='{css_contenedor_col}'><span style='{css_titulo_col}'>C (Protection Leg)</span><span style='{css_valor_col}'>{C:.6f}</span></div>", unsafe_allow_html=True)
    with c4r:
        st.markdown(f"<div style='{css_contenedor_col}'><span style='{css_titulo_col}'>E(T) — Pérdida esperada</span><span style='{css_valor_col}'>{E_T:.6f}</span></div>", unsafe_allow_html=True)

    separador()

    with paso_a_paso():
        st.latex(
            rf"A = {A:.8f}, \quad B = {B:.8f}, \quad C = {C:.8f}"
        )
        st.latex(
            rf"s = \frac{{C}}{{A+B}} = \frac{{{C:.8f}}}{{{A+B:.8f}}} = {_res['spread']:.8f}"
        )
        themed_success(
            f"<div style='{css_paso}'>"
            f"<span style='{math_style}'>s</span> = {_res['spread']:.6f} = "
            f"{spread_bps:.1f} puntos base"
            f"</div>"
        )

    separador()

    # IMPORTANTE: Se agregó unsafe_allow_html=True para que tome el <sub>
    st.markdown(f"#### Pérdida esperada del tranche {_tranche_nombre} por periodo E<sub>j</sub> (promedio sobre F)", unsafe_allow_html=True)
    
    E_all = _res["E_all"]
    taus_full = [0.0] + _periodos
    E_mean = [(W_NODES * E_all[j]).sum() for j in range(len(taus_full))]

    fig_E = go.Figure()
    fig_E.add_trace(go.Scatter(
        x=taus_full, y=E_mean,
        mode="lines+markers",
        line=dict(color=plotly_colors()[0], width=2.5),
        marker=dict(size=8),
        name="E<sub>j</sub> (∫ sobre F)", # Plotly soporta HTML básico aquí
    ))
    fig_E.add_hline(y=1.0, line_dash="dash", line_color="gray", line_width=1,
                    annotation_text="Sin pérdida (E=1)")
    fig_E = apply_plotly_theme(fig_E)
    fig_E.update_layout(
        **plotly_theme(),
        title=f"Pérdida esperada del tranche {_tranche_nombre} por periodo",
        xaxis_title="Tiempo (años)",
        yaxis_title="E<sub>j</sub> (fracción del tranche expuesto)", # Plotly soporta HTML básico aquí
        height=360,
    )
    st.plotly_chart(fig_E, use_container_width=True)

    themed_info(
        f"E<sub>j</sub> representa la **fracción del tranche que aún está en riesgo** en el periodo j. "
        f"Comienza en 1.0 (sin pérdidas) y cae conforme aumentan los defaults esperados. "
        f"Al vencimiento E(T) = **{E_T:.6f}**, lo que significa que en promedio el "
        f"tranche pierde el {(1-E_T)*100:.4f}% de su valor nominal."
    )


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — SENSIBILIDAD
# ─────────────────────────────────────────────────────────────────────────────
with tab_sens:
    st.markdown(f"### Análisis de sensibilidad del spread — Tranche {_tranche_nombre}")
    themed_info(
        "¿Cómo varía el spread cuando cambia un parámetro manteniendo el resto constante? "
        "Esto reproduce la intuición de las griegas de crédito para tranches CDO."
    )

    separador()

    s_col1, s_col2 = st.columns(2)

    with s_col1:
        st.markdown("#### Spread vs Correlación (ρ)")
        rhos = np.linspace(0.01, 0.99, 60)
        spreads_rho = []
        for r_i in rhos:
            ri = valuar_tranche(_h, r_i, _n, _RR, _aL, _aH, _rf, _periodos, F_NODES, W_NODES)
            spreads_rho.append(ri["spread_bps"])

        fig_rho = go.Figure()
        fig_rho.add_trace(go.Scatter(
            x=rhos * 100, y=spreads_rho,
            mode="lines",
            line=dict(color=plotly_colors()[0], width=2.5),
        ))
        fig_rho.add_vline(x=_rho * 100, line_dash="dash",
                          line_color=plotly_colors()[2], line_width=1.5,
                          annotation_text=f"ρ actual = {_rho:.0%}")
        fig_rho = apply_plotly_theme(fig_rho)
        fig_rho.update_layout(
            **plotly_theme(),
            title=f"Spread vs ρ — {_tranche_nombre}",
            xaxis_title="Correlación ρ (%)",
            yaxis_title="Spread (bps)",
            height=320,
        )
        st.plotly_chart(fig_rho, use_container_width=True)

    with s_col2:
        st.markdown("#### Spread vs Hazard Rate (h)")
        hs = np.linspace(0.0001, 0.05, 60)
        spreads_h = []
        for h_i in hs:
            ri = valuar_tranche(h_i, _rho, _n, _RR, _aL, _aH, _rf, _periodos, F_NODES, W_NODES)
            spreads_h.append(ri["spread_bps"])

        fig_h = go.Figure()
        fig_h.add_trace(go.Scatter(
            x=hs * 100, y=spreads_h,
            mode="lines",
            line=dict(color=plotly_colors()[1], width=2.5),
        ))
        fig_h.add_vline(x=_h * 100, line_dash="dash",
                        line_color=plotly_colors()[3], line_width=1.5,
                        annotation_text=f"h actual = {_h:.4%}")
        fig_h = apply_plotly_theme(fig_h)
        fig_h.update_layout(
            **plotly_theme(),
            title=f"Spread vs h — {_tranche_nombre}",
            xaxis_title="Hazard Rate h (%)",
            yaxis_title="Spread (bps)",
            height=320,
        )
        st.plotly_chart(fig_h, use_container_width=True)

    separador()

    s_col3, s_col4 = st.columns(2)

    with s_col3:
        st.markdown("#### Spread vs Recovery Rate (RR)")
        RRs = np.linspace(0.01, 0.90, 50)
        spreads_RR = []
        for rr_i in RRs:
            try:
                ri = valuar_tranche(_h, _rho, _n, rr_i, _aL, _aH, _rf, _periodos, F_NODES, W_NODES)
                spreads_RR.append(ri["spread_bps"])
            except Exception:
                spreads_RR.append(np.nan)

        fig_RR = go.Figure()
        fig_RR.add_trace(go.Scatter(
            x=RRs * 100, y=spreads_RR,
            mode="lines",
            line=dict(color=plotly_colors()[2], width=2.5),
        ))
        fig_RR.add_vline(x=_RR * 100, line_dash="dash",
                         line_color=plotly_colors()[0], line_width=1.5,
                         annotation_text=f"RR actual = {_RR:.0%}")
        fig_RR = apply_plotly_theme(fig_RR)
        fig_RR.update_layout(
            **plotly_theme(),
            title=f"Spread vs RR — {_tranche_nombre}",
            xaxis_title="Recovery Rate RR (%)",
            yaxis_title="Spread (bps)",
            height=320,
        )
        st.plotly_chart(fig_RR, use_container_width=True)

    with s_col4:
        st.markdown("#### Mapa de calor: Spread vs h × ρ")
        h_grid  = np.linspace(0.001, 0.03, 20)
        rho_grid = np.linspace(0.01, 0.80, 20)
        Z_heat  = np.zeros((len(rho_grid), len(h_grid)))
        for i, r_i in enumerate(rho_grid):
            for j, h_i in enumerate(h_grid):
                try:
                    ri = valuar_tranche(h_i, r_i, _n, _RR, _aL, _aH, _rf, _periodos, F_NODES, W_NODES)
                    Z_heat[i, j] = ri["spread_bps"] if not np.isnan(ri["spread_bps"]) else 0
                except Exception:
                    Z_heat[i, j] = 0

        fig_heat = go.Figure(go.Heatmap(
            x=h_grid * 100,
            y=rho_grid * 100,
            z=Z_heat,
            colorscale="RdYlGn_r",
            colorbar=dict(title="Spread (bps)"),
        ))
        fig_heat = apply_plotly_theme(fig_heat)
        fig_heat.update_layout(
            **plotly_theme(),
            title=f"Spread (bps): h vs ρ — {_tranche_nombre}",
            xaxis_title="Hazard Rate h (%)",
            yaxis_title="Correlación ρ (%)",
            height=320,
        )
        st.plotly_chart(fig_heat, use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — TODOS LOS TRANCHES
# ─────────────────────────────────────────────────────────────────────────────
with tab_todos:
    st.markdown("### Comparativa: todos los tranches del CDO")
    themed_info(
        "Con los parámetros del portafolio definidos en la pestaña **Parámetros**, "
        "se valúan simultáneamente **todos los tranches** de la estructura estándar. "
        "Esto permite comparar el spread de cada tranche y entender cómo la subordinación "
        "protege a los tranches senior."
    )

    separador()

    resultados_todos = {}
    for nombre, aL_i, aH_i in TRANCHES_DEFAULT:
        try:
            ri = valuar_tranche(_h, _rho, _n, _RR, aL_i, aH_i, _rf, _periodos, F_NODES, W_NODES)
            resultados_todos[nombre] = ri
        except Exception as e:
            resultados_todos[nombre] = {"spread_bps": np.nan, "E_T": np.nan, "A": np.nan, "B": np.nan, "C": np.nan}

    filas = []
    for nombre, (n2, aL_i, aH_i) in zip([t[0] for t in TRANCHES_DEFAULT], TRANCHES_DEFAULT):
        ri = resultados_todos[nombre]
        _n_L = aL_i * _n / _LGD if _LGD > 0 else 0
        _n_H = aH_i * _n / _LGD if _LGD > 0 else 0
        filas.append({
            "Tranche":        nombre,
            "α_L":            f"{aL_i:.0%}",
            "α_H":            f"{aH_i:.0%}",
            "n_L":            f"{_n_L:.2f}",
            "n_H":            f"{_n_H:.2f}",
            "E(T)":           f"{ri['E_T']:.6f}" if not np.isnan(ri.get('E_T', np.nan)) else "—",
            "A":              f"{ri['A']:.6f}"  if not np.isnan(ri.get('A',   np.nan)) else "—",
            "B":              f"{ri['B']:.6f}"  if not np.isnan(ri.get('B',   np.nan)) else "—",
            "C":              f"{ri['C']:.6f}"  if not np.isnan(ri.get('C',   np.nan)) else "—",
            "Spread (bps)":   f"{ri['spread_bps']:.1f}" if not np.isnan(ri.get('spread_bps', np.nan)) else "—",
        })
    df_todos = pd.DataFrame(filas)
    st.dataframe(df_todos, use_container_width=True, hide_index=True)

    separador()

    nombres_t  = [t[0] for t in TRANCHES_DEFAULT]
    spreads_t  = [resultados_todos[n]["spread_bps"] for n in nombres_t]
    colores_t  = [TRANCHE_COLORES.get(n, "#999") for n in nombres_t]

    fig_bar = go.Figure()
    for nom, sp, col in zip(nombres_t, spreads_t, colores_t):
        fig_bar.add_trace(go.Bar(
            x=[nom], y=[sp if not np.isnan(sp) else 0],
            name=nom,
            marker_color=col,
            text=[f"{sp:.1f}" if not np.isnan(sp) else "N/A"],
            textposition="outside",
        ))
    fig_bar = apply_plotly_theme(fig_bar)
    fig_bar.update_layout(
        **plotly_theme(),
        title=f"Spread por Tranche (h={_h:.4%}, ρ={_rho:.0%}, RR={_RR:.0%})",
        xaxis_title="Tranche",
        yaxis_title="Spread (bps)",
        showlegend=False,
        height=420,
    )
    st.plotly_chart(fig_bar, use_container_width=True)

    separador()

    st.markdown("#### Diagrama de cascada de pérdidas (Waterfall)")
    themed_info(
        "Las pérdidas se absorben de abajo hacia arriba. El tranche **Equity (0–3%)** "
        "es el primero en perder y exige el mayor spread. El tranche **Senior (22–100%)** "
        "solo pierde en escenarios de colapso masivo y tiene el menor spread."
    )

    fig_wf = go.Figure()
    for idx, (nom, aL_i, aH_i) in enumerate(TRANCHES_DEFAULT):
        ancho = aH_i - aL_i
        sp_i  = resultados_todos[nom]["spread_bps"]
        label = f"{nom}<br>{sp_i:.0f} bps" if not np.isnan(sp_i) else f"{nom}<br>N/A"
        fig_wf.add_trace(go.Bar(
            x=[ancho * 100],
            y=[label],
            orientation="h",
            name=nom,
            marker_color=TRANCHE_COLORES.get(nom, "#999"),
            text=[f"{sp_i:.0f} bps" if not np.isnan(sp_i) else "N/A"],
            textposition="inside",
            insidetextanchor="middle",
        ))
    fig_wf = apply_plotly_theme(fig_wf)
    fig_wf.update_layout(
        **plotly_theme(),
        title="Cascada de Tranches (ancho = % del portafolio)",
        xaxis_title="% del Portafolio",
        barmode="stack",
        showlegend=False,
        height=300,
    )
    st.plotly_chart(fig_wf, use_container_width=True)

    separador()

    max_tr   = max(resultados_todos, key=lambda k: resultados_todos[k].get("spread_bps", -1))
    min_tr   = min(resultados_todos, key=lambda k: resultados_todos[k].get("spread_bps", float("inf")))
    max_bps  = resultados_todos[max_tr]["spread_bps"]
    min_bps  = resultados_todos[min_tr]["spread_bps"]

    themed_info(
        f"**Tranche más riesgoso:** {max_tr} con spread de **{max_bps:.1f} bps** — "
        f"absorbe las primeras pérdidas del portafolio. "
        f"**Tranche más seguro (menor spread):** {min_tr} con **{min_bps:.1f} bps** — "
        f"protegido por toda la subordinación de los tranches inferiores."
    )

    separador()
    themed_info(
        "**Referencia:** Hull, J. (2018). **Options, Futures, and Other Derivatives**, "
        "9ª ed., Capítulo 25 — Credit Derivatives. Ecuación 25.12."
    )


# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — EXPORTAR A EXCEL
# ─────────────────────────────────────────────────────────────────────────────
with tab_export:
    st.markdown("### Exportar Arquitectura a Excel")
    themed_info(
        "Extrae la memoria completa de la valuación CDO a una hoja de Excel nativa con "
        "todas las hojas organizadas: parámetros, metodología paso a paso, resultados del "
        "tranche seleccionado, pérdidas esperadas por periodo y comparativa de tranches. "
        "Perfecto para auditar, presentar o continuar el análisis sin Python."
    )

    separador()

    # ── Vista previa del contenido ────────────────────────────────────────────
    col_p1, col_p2, col_p3, col_p4 = st.columns(4)
    col_p1.markdown("**Hoja 1**\nParámetros del portafolio y del tranche")
    col_p2.markdown("**Hoja 2**\nMetodología — fórmulas y valores calculados paso a paso")
    col_p3.markdown("**Hoja 3**\nResultados — A, B, C, spread, E(T)")
    col_p4.markdown("**Hoja 4**\nComparativa de todos los tranches del CDO")

    separador()

    if st.button("Generar Excel de CDO Tranches", use_container_width=True, key="btn_xls_cdo"):
        try:
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # ── Paleta de colores ─────────────────────────────────────────────
            HDR    = Font(bold=True, color="FFFFFF", size=10)
            HDR_SM = Font(bold=True, color="FFFFFF", size=9)
            FIL_AZ = PatternFill("solid", start_color="1F4E79")   # azul oscuro — títulos principales
            FIL_AZ2= PatternFill("solid", start_color="2E75B6")   # azul medio — subtítulos
            FIL_GR = PatternFill("solid", start_color="F2F2F2")   # gris claro — filas alternas
            FIL_VD = PatternFill("solid", start_color="375623")   # verde — resultado destacado
            FIL_VD2= PatternFill("solid", start_color="E2EFDA")   # verde claro — dato ok
            FIL_AM = PatternFill("solid", start_color="FFF2CC")   # amarillo — advertencia
            CTR    = Alignment(horizontal="center", vertical="center")
            LEFT   = Alignment(horizontal="left",   vertical="center")
            RIGHT  = Alignment(horizontal="right",  vertical="center")

            thin = Side(style="thin", color="D6DCE4")
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

            def hdr(ws, r, c, v, fill=FIL_AZ, font=None, align=CTR):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font  = font or HDR
                cell.fill  = fill
                cell.alignment = align
                cell.border = brd
                return cell

            def dat(ws, r, c, v, fill=None, fmt=None, align=RIGHT, bold=False):
                cell = ws.cell(row=r, column=c, value=v)
                if fill:  cell.fill  = fill
                if fmt:   cell.number_format = fmt
                cell.alignment = align
                cell.border = brd
                if bold: cell.font = Font(bold=True, size=10)
                return cell

            def kv_row(ws, r, label, value, fill_val=None, fmt=None):
                """Fila clave–valor: col A label, col B valor."""
                c1 = ws.cell(row=r, column=1, value=label)
                c1.font = Font(bold=True, size=10); c1.alignment = LEFT; c1.border = brd
                c2 = ws.cell(row=r, column=2, value=value)
                c2.alignment = RIGHT; c2.border = brd
                if fill_val: c2.fill = fill_val
                if fmt: c2.number_format = fmt
                return r + 1

            def section_title(ws, r, text, ncols=8, fill=FIL_AZ2):
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
                cell = ws.cell(row=r, column=1, value=text)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = fill; cell.alignment = LEFT
                return r + 1

            # ═══════════════════════════════════════════════════════════════════
            # HOJA 1 — PARÁMETROS
            # ═══════════════════════════════════════════════════════════════════
            ws1 = wb.active
            ws1.title = "1 Parámetros"
            ws1.column_dimensions["A"].width = 34
            ws1.column_dimensions["B"].width = 24

            # Título principal
            ws1.merge_cells("A1:B1")
            tc = ws1.cell(row=1, column=1, value="CDO TRANCHE VALUATION — Parámetros de Entrada")
            tc.font = Font(bold=True, color="FFFFFF", size=13)
            tc.fill = FIL_AZ; tc.alignment = CTR

            ws1.row_dimensions[1].height = 22
            r = 2

            # Bloque 1: Portafolio
            r = section_title(ws1, r, "1.1  Portafolio de Referencia", ncols=2)
            r = kv_row(ws1, r, "Valor del portafolio ($)",          _VP,          fmt='#,##0.00')
            r = kv_row(ws1, r, "Número de entidades (n)",           _n,           fmt='0')
            r = kv_row(ws1, r, "Vencimiento del CDS (T, años)",     _T,           fmt='0.00')
            r = kv_row(ws1, r, "Periodo de pago de primas (Δt)",    _dt,          fmt='0.00')
            r = kv_row(ws1, r, "Número de periodos de pago",        len(_periodos), fmt='0')
            periodos_str = ", ".join([f"{p:.2f}" for p in _periodos])
            r = kv_row(ws1, r, "Periodos τ (años)",                 periodos_str)
            r += 1

            # Bloque 2: Crédito
            r = section_title(ws1, r, "1.2  Parámetros de Crédito", ncols=2)
            r = kv_row(ws1, r, "Hazard rate (h)",                    _h,           fmt='0.000000')
            r = kv_row(ws1, r, "Hazard rate (h) en %",               _h * 100,     fmt='0.0000"%"')
            r = kv_row(ws1, r, "Recovery Rate (RR)",                 _RR,          fmt='0.00%')
            r = kv_row(ws1, r, "Loss Given Default (LGD = 1−RR)",    _LGD,         fmt='0.00%')
            r = kv_row(ws1, r, "Correlación (ρ)",                    _rho,         fmt='0.00%')
            r = kv_row(ws1, r, "Tasa libre de riesgo (r)",           _rf,          fmt='0.00%')
            r = kv_row(ws1, r, "Puntos de cuadratura Gauss-Hermite (M)", _M,       fmt='0')
            r += 1

            # Bloque 3: Tranche
            r = section_title(ws1, r, "1.3  Tranche Analizado", ncols=2)
            r = kv_row(ws1, r, "Nombre del tranche",                 _tranche_nombre)
            r = kv_row(ws1, r, "Attachment point (α_L)",             _aL,          fmt='0.00%')
            r = kv_row(ws1, r, "Detachment point (α_H)",             _aH,          fmt='0.00%')
            r = kv_row(ws1, r, "Ancho del tranche (α_H − α_L)",      _aH - _aL,    fmt='0.00%')
            _n_L_ex = _aL * _n / _LGD if _LGD > 0 else 0
            _n_H_ex = _aH * _n / _LGD if _LGD > 0 else 0
            r = kv_row(ws1, r, "n_L = α_L · n / LGD",               _n_L_ex,      fmt='0.0000')
            r = kv_row(ws1, r, "n_H = α_H · n / LGD",               _n_H_ex,      fmt='0.0000')
            r = kv_row(ws1, r, "m(n_L) = ⌈n_L⌉",                   int(np.ceil(_n_L_ex)), fmt='0')
            r = kv_row(ws1, r, "m(n_H) = ⌈n_H⌉",                   int(np.ceil(_n_H_ex)), fmt='0')
            r = kv_row(ws1, r, "Valor nominal del tranche ($)",      (_aH - _aL) * _VP,  fmt='#,##0.00')

            # ═══════════════════════════════════════════════════════════════════
            # HOJA 2 — METODOLOGÍA
            # ═══════════════════════════════════════════════════════════════════
            ws2 = wb.create_sheet("2 Metodología")
            ws2.column_dimensions["A"].width = 38
            ws2.column_dimensions["B"].width = 22
            ws2.column_dimensions["C"].width = 22
            ws2.column_dimensions["D"].width = 22

            ws2.merge_cells("A1:D1")
            tc2 = ws2.cell(row=1, column=1,
                           value="Metodología: Modelo de un Factor Gaussiano (Hull, 9ª ed., ec. 25.12)")
            tc2.font = Font(bold=True, color="FFFFFF", size=13)
            tc2.fill = FIL_AZ; tc2.alignment = CTR
            ws2.row_dimensions[1].height = 22
            r = 2

            # Paso 1: Q(t)
            r = section_title(ws2, r, "Paso 1 — Probabilidad de default Q(t) = 1 − exp(−h·t)", ncols=4)
            hdr(ws2, r, 1, "Periodo τ",       fill=FIL_AZ2)
            hdr(ws2, r, 2, "Q(τ)",             fill=FIL_AZ2)
            hdr(ws2, r, 3, "Q(τ) en %",        fill=FIL_AZ2)
            hdr(ws2, r, 4, "N⁻¹[Q(τ)]",       fill=FIL_AZ2)
            r += 1
            for t_j in _periodos:
                Qt_j   = 1.0 - np.exp(-_h * t_j)
                inv_Qt = norm.ppf(Qt_j)
                fill_row = FIL_GR if r % 2 == 0 else None
                dat(ws2, r, 1, t_j,       fill=fill_row, fmt='0.00')
                dat(ws2, r, 2, Qt_j,      fill=fill_row, fmt='0.00000000')
                dat(ws2, r, 3, Qt_j*100,  fill=fill_row, fmt='0.0000"%"')
                dat(ws2, r, 4, inv_Qt,    fill=fill_row, fmt='0.000000')
                r += 1
            r += 1

            # Paso 2: Q(t|F) para F=0
            r = section_title(ws2, r,
                "Paso 2 — Prob. condicional Q(t|F=0) = N([N⁻¹(Q) − √ρ·F] / √(1−ρ))", ncols=4)
            hdr(ws2, r, 1, "Periodo τ",    fill=FIL_AZ2)
            hdr(ws2, r, 2, "Q(τ)",         fill=FIL_AZ2)
            hdr(ws2, r, 3, "arg(F=0)",     fill=FIL_AZ2)
            hdr(ws2, r, 4, "Q(τ|F=0)",     fill=FIL_AZ2)
            r += 1
            for t_j in _periodos:
                Qt_j   = 1.0 - np.exp(-_h * t_j)
                inv_Qt = norm.ppf(Qt_j)
                arg_0  = (inv_Qt - np.sqrt(_rho) * 0.0) / np.sqrt(1.0 - _rho)
                Q_0    = float(norm.cdf(arg_0))
                fill_row = FIL_GR if r % 2 == 0 else None
                dat(ws2, r, 1, t_j,    fill=fill_row, fmt='0.00')
                dat(ws2, r, 2, Qt_j,   fill=fill_row, fmt='0.00000000')
                dat(ws2, r, 3, arg_0,  fill=fill_row, fmt='0.000000')
                dat(ws2, r, 4, Q_0,    fill=fill_row, fmt='0.00000000')
                r += 1
            r += 1

            # Paso 3: Cuadratura — primeros 10 nodos
            r = section_title(ws2, r,
                f"Paso 3 — Cuadratura Gauss-Hermite (M={_M} nodos; se muestran los primeros 10)", ncols=4)
            hdr(ws2, r, 1, "k",      fill=FIL_AZ2)
            hdr(ws2, r, 2, "F_k",    fill=FIL_AZ2)
            hdr(ws2, r, 3, "w_k",    fill=FIL_AZ2)
            hdr(ws2, r, 4, "F_k²/2", fill=FIL_AZ2)
            r += 1
            _show_nodes = min(10, _M)
            for ki in range(_show_nodes):
                fill_row = FIL_GR if r % 2 == 0 else None
                dat(ws2, r, 1, ki + 1,                    fill=fill_row, fmt='0')
                dat(ws2, r, 2, float(F_NODES[ki]),        fill=fill_row, fmt='0.000000')
                dat(ws2, r, 3, float(W_NODES[ki]),        fill=fill_row, fmt='0.00000000')
                dat(ws2, r, 4, float(F_NODES[ki]**2/2),  fill=fill_row, fmt='0.000000')
                r += 1
            # Verificación suma pesos
            ws2.cell(row=r, column=1,
                     value=f"Suma de todos los {_M} pesos (debe ≈ 1)").font = Font(italic=True, size=9)
            dat(ws2, r, 2, float(W_NODES.sum()), fill=FIL_VD2, fmt='0.00000000')
            r += 2

            # Paso 4: E_j(F=0) por periodo
            r = section_title(ws2, r,
                "Paso 4 — Pérdida esperada E_j(F=0) del tranche por periodo", ncols=4)
            hdr(ws2, r, 1, "Periodo τ",    fill=FIL_AZ2)
            hdr(ws2, r, 2, "E_j(F=0)",     fill=FIL_AZ2)
            hdr(ws2, r, 3, "ΔE_j = E_{j-1}−E_j", fill=FIL_AZ2)
            hdr(ws2, r, 4, "v(τ) = e^{−r·τ}", fill=FIL_AZ2)
            r += 1
            _taus_full2 = [0.0] + _periodos
            _E_all2 = _res["E_all"]
            _E_mean2 = [(W_NODES * _E_all2[j]).sum() for j in range(len(_taus_full2))]
            for j_idx, t_j in enumerate(_taus_full2):
                Ej_val  = _E_mean2[j_idx]
                dEj     = (_E_mean2[j_idx - 1] - Ej_val) if j_idx > 0 else 0.0
                v_j     = float(np.exp(-_rf * t_j))
                fill_row = FIL_GR if r % 2 == 0 else None
                dat(ws2, r, 1, t_j,    fill=fill_row, fmt='0.00')
                dat(ws2, r, 2, float(Ej_val), fill=fill_row, fmt='0.00000000')
                dat(ws2, r, 3, float(dEj),    fill=fill_row, fmt='0.00000000')
                dat(ws2, r, 4, v_j,           fill=fill_row, fmt='0.00000000')
                r += 1

            # ═══════════════════════════════════════════════════════════════════
            # HOJA 3 — RESULTADOS
            # ═══════════════════════════════════════════════════════════════════
            ws3 = wb.create_sheet("3 Resultados")
            ws3.column_dimensions["A"].width = 38
            ws3.column_dimensions["B"].width = 26

            ws3.merge_cells("A1:B1")
            tc3 = ws3.cell(row=1, column=1,
                           value=f"Resultados — Tranche {_tranche_nombre}  [{_aL:.0%}–{_aH:.0%}]")
            tc3.font = Font(bold=True, color="FFFFFF", size=13)
            tc3.fill = FIL_VD; tc3.alignment = CTR
            ws3.row_dimensions[1].height = 22
            r = 2

            # Resultado principal — spread destacado
            r = section_title(ws3, r, "Resultado Principal", ncols=2, fill=FIL_VD)
            r = kv_row(ws3, r, f"SPREAD del tranche {_tranche_nombre}",
                       _res["spread_bps"],
                       fill_val=FIL_VD2, fmt='0.0000 "bps"')
            r = kv_row(ws3, r, "Spread (decimal)",
                       _res["spread"],
                       fill_val=FIL_VD2, fmt='0.00000000')
            r += 1

            # Componentes
            r = section_title(ws3, r, "Componentes de la Fee Leg y Protection Leg", ncols=2)
            r = kv_row(ws3, r, "A  (Fee Leg principal)",          _res["A"],  fmt='0.00000000')
            r = kv_row(ws3, r, "B  (Fee Leg accrual)",            _res["B"],  fmt='0.00000000')
            r = kv_row(ws3, r, "A + B  (denominador del spread)", _res["A"] + _res["B"], fmt='0.00000000')
            r = kv_row(ws3, r, "C  (Protection Leg)",             _res["C"],  fmt='0.00000000')
            r = kv_row(ws3, r, "s = C / (A+B)",                   _res["spread"], fmt='0.00000000')
            r += 1

            # Pérdida esperada
            r = section_title(ws3, r, "Pérdida Esperada", ncols=2)
            E_T = _res["E_T"]
            r = kv_row(ws3, r, "E(T) — fracción del tranche en riesgo al vto.", E_T,
                       fmt='0.00000000')
            r = kv_row(ws3, r, "Pérdida media del tranche (1 − E(T))",          1 - E_T,
                       fmt='0.00%')
            r += 1

            # Tabla E_j por periodo (hoja 3 también)
            r = section_title(ws3, r,
                f"Pérdida esperada E_j por periodo — tranche {_tranche_nombre} (∫ sobre F)", ncols=2)
            hdr(ws3, r, 1, "Periodo τ (años)", fill=FIL_AZ2)
            hdr(ws3, r, 2, "E_j (promedio sobre F)", fill=FIL_AZ2)
            r += 1
            _taus_full3 = [0.0] + _periodos
            _E_all3     = _res["E_all"]
            _E_mean3    = [(W_NODES * _E_all3[j]).sum() for j in range(len(_taus_full3))]
            for j_idx, t_j in enumerate(_taus_full3):
                fill_row = FIL_GR if r % 2 == 0 else None
                dat(ws3, r, 1, t_j,               fill=fill_row, fmt='0.00')
                dat(ws3, r, 2, float(_E_mean3[j_idx]), fill=fill_row, fmt='0.00000000')
                r += 1

            # ═══════════════════════════════════════════════════════════════════
            # HOJA 4 — COMPARATIVA DE TRANCHES
            # ═══════════════════════════════════════════════════════════════════
            ws4 = wb.create_sheet("4 Comparativa Tranches")

            _col_widths4 = [14, 10, 10, 12, 12, 16, 16, 16, 16, 16]
            _headers4    = ["Tranche", "α_L", "α_H", "n_L", "n_H",
                            "A", "B", "C", "E(T)", "Spread (bps)"]
            for ci, (h_txt, cw) in enumerate(zip(_headers4, _col_widths4), 1):
                ws4.column_dimensions[get_column_letter(ci)].width = cw

            ws4.merge_cells(f"A1:{get_column_letter(len(_headers4))}1")
            tc4 = ws4.cell(row=1, column=1,
                           value=f"Comparativa de Tranches  (h={_h:.4%} | ρ={_rho:.0%} | RR={_RR:.0%})")
            tc4.font = Font(bold=True, color="FFFFFF", size=13)
            tc4.fill = FIL_AZ; tc4.alignment = CTR
            ws4.row_dimensions[1].height = 22

            # Cabecera
            for ci, h_txt in enumerate(_headers4, 1):
                hdr(ws4, 2, ci, h_txt, fill=FIL_AZ2)

            # Calcular todos los tranches
            _resultados_xls = {}
            for nom_t, aL_t, aH_t in TRANCHES_DEFAULT:
                try:
                    ri_t = valuar_tranche(
                        _h, _rho, _n, _RR, aL_t, aH_t, _rf, _periodos, F_NODES, W_NODES
                    )
                    _resultados_xls[nom_t] = ri_t
                except Exception:
                    _resultados_xls[nom_t] = {
                        "spread_bps": np.nan, "E_T": np.nan,
                        "A": np.nan, "B": np.nan, "C": np.nan,
                    }

            r = 3
            for idx, (nom_t, aL_t, aH_t) in enumerate(TRANCHES_DEFAULT):
                ri_t   = _resultados_xls[nom_t]
                n_L_t  = aL_t * _n / _LGD if _LGD > 0 else 0
                n_H_t  = aH_t * _n / _LGD if _LGD > 0 else 0
                es_sel = (nom_t == _tranche_nombre)
                fill_row = FIL_VD2 if es_sel else (FIL_GR if idx % 2 == 0 else None)

                dat(ws4, r, 1,  nom_t,                        fill=fill_row, align=LEFT, bold=es_sel)
                dat(ws4, r, 2,  aL_t,                         fill=fill_row, fmt='0%')
                dat(ws4, r, 3,  aH_t,                         fill=fill_row, fmt='0%')
                dat(ws4, r, 4,  n_L_t,                        fill=fill_row, fmt='0.0000')
                dat(ws4, r, 5,  n_H_t,                        fill=fill_row, fmt='0.0000')
                dat(ws4, r, 6,  float(ri_t["A"])   if not np.isnan(ri_t["A"])   else None,
                    fill=fill_row, fmt='0.00000000')
                dat(ws4, r, 7,  float(ri_t["B"])   if not np.isnan(ri_t["B"])   else None,
                    fill=fill_row, fmt='0.00000000')
                dat(ws4, r, 8,  float(ri_t["C"])   if not np.isnan(ri_t["C"])   else None,
                    fill=fill_row, fmt='0.00000000')
                dat(ws4, r, 9,  float(ri_t["E_T"]) if not np.isnan(ri_t["E_T"]) else None,
                    fill=fill_row, fmt='0.00000000')
                dat(ws4, r, 10, float(ri_t["spread_bps"]) if not np.isnan(ri_t["spread_bps"]) else None,
                    fill=FIL_VD2 if es_sel else fill_row,
                    fmt='0.00 "bps"', bold=es_sel)
                r += 1

            # Fila de referencia
            r += 1
            ws4.cell(row=r, column=1,
                     value="* Tranche seleccionado resaltado en verde.").font = Font(italic=True, size=9, color="375623")
            r += 1
            ws4.cell(row=r, column=1,
                     value="Referencia: Hull, J. (2018). Options, Futures, and Other Derivatives, 9ª ed., Cap. 25.").font = \
                Font(italic=True, size=9, color="777777")

            # ── Guardar y ofrecer descarga ────────────────────────────────────
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.session_state["cdo_excel"] = buf.getvalue()
            themed_success(
                f"✅ Excel generado — Tranche {_tranche_nombre} · "
                f"Spread: {_res['spread_bps']:.1f} bps · "
                f"4 hojas con {len(_periodos)} periodos"
            )

        except Exception as e:
            themed_error(f"❌ Error al generar el Excel: {e}")

    if "cdo_excel" in st.session_state:
        st.download_button(
            "⬇️ Descargar CDO Tranches (.xlsx)",
            data=st.session_state["cdo_excel"],
            file_name=f"CDO_{_tranche_nombre.replace('–', '-').replace('%', 'pct')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    separador()
    themed_info(
        "**Contenido del Excel generado:**\n\n"
        "- **Hoja 1 — Parámetros:** portafolio, crédito y tranche (attachment, detachment, n<sub>L</sub>, n<sub>H</sub>)\n"
        "- **Hoja 2 — Metodología:** Q(t) por periodo, Q(t|F=0), nodos y pesos de Gauss-Hermite, E<sub>j</sub>(F=0)\n"
        "- **Hoja 3 — Resultados:** spread, A, B, C, E(T), pérdida esperada y E<sub>j</sub> integrado por periodo\n"
        "- **Hoja 4 — Comparativa:** todos los tranches del CDO con sus métricas, tranche seleccionado resaltado\n\n"
        " Referencia: Hull, J. (2018). *Options, Futures, and Other Derivatives*, 9ª ed., Cap. 25, ec. 25.12."
    )